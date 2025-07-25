import os
import openpyxl
import requests


def check_document(document_id, base_url, api_key):
    """Query Asset Suite 9 for the given document ID.

    Parameters
    ----------
    document_id : str
        ID or name of the document to look up.
    base_url : str
        Base URL for the Asset Suite 9 API. Something like
        ``https://example.com/api``.
    api_key : str
        API key or token used for authorization.
    Returns
    -------
    bool
        True if the document exists, False if not found.
    """
    url = f"{base_url}/documents/{document_id}"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Accept": "application/json",
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return True
    elif response.status_code == 404:
        return False
    else:
        response.raise_for_status()
        return False


def process_excel(in_path, out_path, base_url, api_key):
    """Read document IDs from the Excel file and check each one."""
    wb = openpyxl.load_workbook(in_path)
    sheet = wb.active

    # Assume first column contains document IDs
    for row in sheet.iter_rows(min_row=2):
        cell = row[0]
        doc_id = str(cell.value).strip() if cell.value is not None else ""
        if not doc_id:
            continue
        try:
            exists = check_document(doc_id, base_url, api_key)
            result = "FOUND" if exists else "MISSING"
        except Exception as exc:
            result = f"ERROR: {exc}"
        # Write result to second column
        sheet.cell(row=cell.row, column=2, value=result)

    wb.save(out_path)


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Check Asset Suite 9 documents from Excel list")
    parser.add_argument("excel", help="Path to the Excel file containing document IDs in column A")
    parser.add_argument("-o", "--output", default="results.xlsx", help="Output Excel file with results")
    parser.add_argument("--url", default=os.environ.get("AS9_URL", ""), help="Base URL for Asset Suite 9 API")
    parser.add_argument("--key", default=os.environ.get("AS9_KEY", ""), help="API key or token for Asset Suite 9")

    args = parser.parse_args()

    if not args.url or not args.key:
        parser.error("API URL and key must be provided via arguments or environment variables")

    process_excel(args.excel, args.output, args.url, args.key)
    print(f"Results written to {args.output}")


if __name__ == "__main__":
    main()
